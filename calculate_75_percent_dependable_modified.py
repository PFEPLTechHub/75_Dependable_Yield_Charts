import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

def _sanitize_sheet_name(name: str) -> str:
    """Sanitize sheet names to meet Excel requirements."""
    name = re.sub(r'[:\\\/\?\*\[\]]', '_', str(name))
    return name[:31]

def seventy_five_dependability(input_excel, output_excel=None, num_years=None, num_days=None):
    """
    Calculate 75% dependable yield from multi-sheet Excel file.
    
    Args:
        input_excel: Path to input Excel with multiple sheets (one per junction)
        output_excel: Path to output Excel (if None, uses input filename + '_75percent_output.xlsx')
        num_years: Number of years to use for calculation (default: None, uses ALL years)
        num_days: Number of days to process (default: None, processes all days)
    """
    
    if output_excel is None:
        base_name = os.path.splitext(input_excel)[0]
        output_excel = f"{base_name}_75percent_output.xlsx"
    
    print(f"\n{'='*80}")
    print(f"75% DEPENDABLE YIELD CALCULATION")
    print(f"{'='*80}")
    print(f"Input file: {input_excel}")
    print(f"Output file: {output_excel}")
    if num_years:
        print(f"Using first {num_years} years for calculation")
    else:
        print(f"Using ALL years available in the data")
    if num_days:
        print(f"Processing only first {num_days} days (for testing)")
    print(f"{'='*80}\n")
    
    # Read all sheets from input Excel
    xl = pd.ExcelFile(input_excel)
    sheet_names = xl.sheet_names
    
    print(f"Found {len(sheet_names)} junctions (sheets):")
    for i, name in enumerate(sheet_names, 1):
        print(f"  {i}. {name}")
    
    # Dictionary to store results
    summary_results = {}
    junction_details = {}
    
    # Process each sheet (junction)
    print(f"\n{'='*80}")
    print("PROCESSING EACH JUNCTION")
    print(f"{'='*80}\n")
    
    for sheet_name in sheet_names:
        print(f"Processing: {sheet_name}")
        
        # Read the sheet (with optional row limit for testing)
        if num_days:
            df = pd.read_excel(xl, sheet_name=sheet_name, nrows=num_days)
        else:
            df = pd.read_excel(xl, sheet_name=sheet_name)
        
        # Get year columns (exclude 'Date')
        year_columns = [col for col in df.columns if col != 'Date']
        
        # Use all years or first N years
        if num_years:
            years_to_use = year_columns[:num_years]
        else:
            years_to_use = year_columns
        
        print(f"  Years used: {years_to_use[0]} to {years_to_use[-1]} ({len(years_to_use)} years)")
        print(f"  Total dates: {len(df)}")
        
        # Select only Date and selected years
        data_years = df[years_to_use]
        
        # Calculate 75% dependable yield (25th percentile) for each row (date)
        dependable_75 = data_years.quantile(q=0.25, axis=1)
        
        # Store results
        summary_results[sheet_name] = dependable_75
        
        # Create detailed DataFrame for this junction
        detail_df = df[['Date'] + years_to_use].copy()
        detail_df['75% Dependable'] = dependable_75
        detail_df['Min'] = data_years.min(axis=1)
        detail_df['Max'] = data_years.max(axis=1)
        detail_df['Mean'] = data_years.mean(axis=1)
        
        junction_details[_sanitize_sheet_name(sheet_name)] = detail_df
        
        print(f"  [OK] Calculated 75% dependable yield")
        print(f"    Range: {dependable_75.min():.4f} to {dependable_75.max():.4f}")
        print()
    
    # Create main summary DataFrame
    print(f"{'='*80}")
    print("CREATING OUTPUT FILE")
    print(f"{'='*80}\n")
    
    summary_df = pd.DataFrame(summary_results)
    summary_df.insert(0, 'Date', df['Date'])
    
    print(f"Summary sheet created:")
    print(f"  Dimensions: {summary_df.shape[0]} rows x {summary_df.shape[1]} columns")
    print(f"  Columns: Date + {len(sheet_names)} junctions")
    
    # Write to Excel with multiple sheets
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        # Write main summary sheet
        summary_df.to_excel(writer, sheet_name='75_Percent_Dependability', index=False)
        
        # Write detailed sheets for each junction
        for sheet_name, detail_df in junction_details.items():
            detail_df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    print(f"\n[OK] Written main summary sheet: '75_Percent_Dependability'")
    print(f"[OK] Written {len(junction_details)} detailed junction sheets")
    
    # Now format the Excel file with styles
    print(f"\n{'='*80}")
    print("APPLYING FORMATTING")
    print(f"{'='*80}\n")
    
    wb = load_workbook(output_excel)
    
    # Format main summary sheet
    if '75_Percent_Dependability' in wb.sheetnames:
        ws = wb['75_Percent_Dependability']
        
        # Header formatting
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set column widths
        ws.column_dimensions['A'].width = 12  # Date column
        for col in range(2, ws.max_column + 1):
            ws.column_dimensions[ws.cell(1, col).column_letter].width = 15
        
        # Freeze first row
        ws.freeze_panes = "A2"
        
        print("[OK] Formatted main summary sheet")
    
    # Format individual junction sheets
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for sheet_name in junction_details.keys():
        if sheet_name not in wb.sheetnames:
            continue
        
        ws = wb[sheet_name]
        
        # Header formatting
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Highlight the "75% Dependable" column
        dep_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(1, col).value and "75%" in str(ws.cell(1, col).value):
                dep_col = col
                break
        
        if dep_col:
            for row in range(2, ws.max_row + 1):
                ws.cell(row, dep_col).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        # Set column widths
        ws.column_dimensions['A'].width = 12
        for col in range(2, ws.max_column + 1):
            ws.column_dimensions[ws.cell(1, col).column_letter].width = 13
        
        # Freeze first row
        ws.freeze_panes = "A2"
    
    print(f"[OK] Formatted {len(junction_details)} junction sheets")
    
    # Save workbook
    wb.save(output_excel)
    
    print(f"\n{'='*80}")
    print("[SUCCESS] COMPLETED SUCCESSFULLY!")
    print(f"{'='*80}")
    print(f"\nOutput saved to: {output_excel}")
    print(f"\nFile contains:")
    print(f"  - Main sheet: '75_Percent_Dependability' (date + all junctions)")
    print(f"  - {len(junction_details)} detailed sheets (one per junction)")
    print(f"\n{'='*80}\n")


# === Main execution ===
if __name__ == "__main__":
    INPUT_PATH = r"ALL MERGED 22 Points_09112025.xlsx"
    
    # Calculate 75% dependable yield using ALL years and ALL dates
    seventy_five_dependability(
        input_excel=INPUT_PATH,
        output_excel="75_Percent_Dependable_FULL.xlsx",
        num_years=None,  # Use ALL years available
        num_days=None    # Process ALL dates (153 days)
    )

